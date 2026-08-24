"""
Gincana Escolar - Sistema de Pontuação
========================================
Flask app mobile-first para avaliação de equipes por professores.
Banco SQLite local (portável).
"""
import os, secrets
from datetime import datetime
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, abort, send_file
from werkzeug.security import generate_password_hash, check_password_hash
import sqlite3
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# On platforms like Render the working dir is read-only; prefer a writable volume.
DATA_DIR = os.environ.get("DATA_DIR", BASE_DIR)
DB_PATH  = os.path.join(DATA_DIR, "gincana.db")

app = Flask(__name__)
# Production-safe: persistent secret key (so sessions survive restarts), override via env.
app.secret_key = os.environ.get(
    "SECRET_KEY",
    "gincana-escolar-prod-7f3a9b2c1d8e4f6a0b5c9d2e7f1a3b8c",
)
app.config["PERMANENT_SESSION_LIFETIME"] = 86400  # 24 h
DEBUG = os.environ.get("FLASK_DEBUG") == "1"

# ---------------------------------------------------------------------------
# DB helpers
# ---------------------------------------------------------------------------
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn

def init_db():
    db = get_db()
    db.executescript("""
    CREATE TABLE IF NOT EXISTS users (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        name       TEXT NOT NULL,
        email      TEXT UNIQUE NOT NULL,
        password   TEXT NOT NULL,
        role       TEXT NOT NULL DEFAULT 'professor',   -- 'admin' | 'professor'
        created_at TEXT DEFAULT (datetime('now'))
    );

    CREATE TABLE IF NOT EXISTS teams (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        name       TEXT NOT NULL UNIQUE,
        color      TEXT DEFAULT '#4F46E5',
        created_at TEXT DEFAULT (datetime('now'))
    );

    CREATE TABLE IF NOT EXISTS categories (
        id    INTEGER PRIMARY KEY AUTOINCREMENT,
        name  TEXT NOT NULL,
        order_idx INTEGER DEFAULT 0
    );

    CREATE TABLE IF NOT EXISTS scores (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        team_id     INTEGER NOT NULL REFERENCES teams(id) ON DELETE CASCADE,
        category_id INTEGER REFERENCES categories(id) ON DELETE SET NULL,
        evaluator_id INTEGER NOT NULL REFERENCES users(id),
        points      REAL NOT NULL,
        note        TEXT DEFAULT '',
        created_at  TEXT DEFAULT (datetime('now')),
        UNIQUE(team_id, category_id, evaluator_id)
    );
    """)

    # Ensure admin user exists
    admin = db.execute("SELECT id FROM users WHERE role='admin'").fetchone()
    if not admin:
        db.execute(
            "INSERT INTO users (name, email, password, role) VALUES (?, ?, ?, ?)",
            ("Coordenador", "admin@gincana.local",
             generate_password_hash("admin123"), "admin"),
        )
    db.commit()
    db.close()

# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------
def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return wrapper

def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        if session.get("role") not in ("admin", "coordinator"):
            abort(403)
        return f(*args, **kwargs)
    return wrapper

def current_user():
    if "user_id" in session:
        db = get_db()
        user = db.execute("SELECT * FROM users WHERE id=?", (session["user_id"],)).fetchone()
        db.close()
        return user
    return None

# ---------------------------------------------------------------------------
# Context processor – gives templates access to current_user / guncana_name
# ---------------------------------------------------------------------------
@app.context_processor
def inject_user():
    return dict(current_user=current_user(), gincana_name="Gincana UP 2026")

# ---------------------------------------------------------------------------
# Public routes
# ---------------------------------------------------------------------------
@app.route("/")
def index():
    if "user_id" in session:
        if session.get("role") in ("admin", "coordinator"):
            return redirect(url_for("admin_dashboard"))
        return redirect(url_for("professor_dashboard"))
    return redirect(url_for("login"))

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        db = get_db()
        user = db.execute("SELECT * FROM users WHERE email=?", (email,)).fetchone()
        db.close()

        if user and check_password_hash(user["password"], password):
            session.permanent = True
            session["user_id"] = user["id"]
            session["role"] = user["role"]
            session["name"] = user["name"]
            if user["role"] in ("admin", "coordinator"):
                return redirect(url_for("admin_dashboard"))
            return redirect(url_for("professor_dashboard"))

        flash("E-mail ou senha incorretos.", "error")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# ---------------------------------------------------------------------------
# Ranking (public)
# ---------------------------------------------------------------------------
@app.route("/ranking")
def ranking():
    db = get_db()
    rows = db.execute("""
        SELECT t.id, t.name, t.color,
               COALESCE(SUM(s.points), 0) AS total,
               COUNT(DISTINCT s.evaluator_id) AS evaluator_count
        FROM teams t
        LEFT JOIN scores s ON s.team_id = t.id
        GROUP BY t.id
        ORDER BY total DESC, t.name ASC
    """).fetchall()
    db.close()
    return render_template("ranking.html", teams=rows, rank=1)

# ---------------------------------------------------------------------------
# Admin routes
# ---------------------------------------------------------------------------
@app.route("/admin")
@admin_required
def admin_dashboard():
    db = get_db()
    teams      = db.execute("SELECT COUNT(*) c FROM teams").fetchone()["c"]
    professors = db.execute("SELECT COUNT(*) c FROM users WHERE role IN ('professor','coordinator')").fetchone()["c"]
    scores     = db.execute("SELECT COUNT(*) c FROM scores").fetchone()["c"]
    db.close()
    return render_template("admin/dashboard.html", teams=teams, professors=professors, scores=scores)

# --- Teams ---
@app.route("/admin/teams", methods=["GET", "POST"])
@admin_required
def admin_teams():
    db = get_db()
    if request.method == "POST":
        name  = request.form.get("name", "").strip()
        color = request.form.get("color", "#4F46E5").strip()
        if not name:
            flash("Nome da equipe é obrigatório.", "error")
        else:
            try:
                db.execute("INSERT INTO teams (name, color) VALUES (?, ?)", (name, color))
                db.commit()
                flash(f"Equipe '{name}' criada!", "success")
            except sqlite3.IntegrityError:
                flash("Já existe uma equipe com esse nome.", "error")
        db.close()
        return redirect(url_for("admin_teams"))

    teams = db.execute("SELECT * FROM teams ORDER BY name").fetchall()
    db.close()
    return render_template("admin/teams.html", teams=teams)

@app.route("/admin/teams/<int:team_id>/delete", methods=["POST"])
@admin_required
def admin_team_delete(team_id):
    db = get_db()
    db.execute("DELETE FROM teams WHERE id=?", (team_id,))
    db.commit()
    db.close()
    flash("Equipe removida.", "success")
    return redirect(url_for("admin_teams"))

# --- Professors ---
@app.route("/admin/professors", methods=["GET", "POST"])
@admin_required
def admin_professors():
    db = get_db()
    if request.method == "POST":
        name     = request.form.get("name", "").strip()
        email    = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "").strip()
        if not name or not email or not password:
            flash("Preencha todos os campos.", "error")
        else:
            try:
                db.execute(
                    "INSERT INTO users (name, email, password, role) VALUES (?, ?, ?, 'professor')",
                    (name, email, generate_password_hash(password)),
                )
                db.commit()
                flash(f"Professor(a) '{name}' cadastrado(a)!", "success")
            except sqlite3.IntegrityError:
                flash("Já existe um cadastro com esse e-mail.", "error")
        db.close()
        return redirect(url_for("admin_professors"))

    profs = db.execute("SELECT id, name, email, role, created_at FROM users WHERE role IN ('professor','coordinator') ORDER BY role DESC, name").fetchall()
    db.close()
    return render_template("admin/professors.html", professors=profs)

@app.route("/admin/professors/<int:prof_id>/promote", methods=["POST"])
@admin_required
def admin_professor_promote(prof_id):
    if session.get("role") != "admin":
        flash("Apenas o coordenador principal pode conceder poderes de coordenador.", "error")
        return redirect(url_for("admin_professors"))
    db = get_db()
    db.execute("UPDATE users SET role='coordinator' WHERE id=?", (prof_id,))
    db.commit()
    user = db.execute("SELECT name FROM users WHERE id=?", (prof_id,)).fetchone()
    db.close()
    if user:
        flash(f"{user['name']} agora é Coordenador(a) e pode acessar o painel e avaliar equipes!", "success")
    return redirect(url_for("admin_professors"))

@app.route("/admin/professors/<int:prof_id>/demote", methods=["POST"])
@admin_required
def admin_professor_demote(prof_id):
    if session.get("role") != "admin":
        flash("Apenas o coordenador principal pode remover poderes de coordenador.", "error")
        return redirect(url_for("admin_professors"))
    db = get_db()
    db.execute("UPDATE users SET role='professor' WHERE id=?", (prof_id,))
    db.commit()
    user = db.execute("SELECT name FROM users WHERE id=?", (prof_id,)).fetchone()
    db.close()
    if user:
        flash(f"{user['name']} voltou a ser apenas Professor(a).", "success")
    return redirect(url_for("admin_professors"))

@app.route("/admin/professors/<int:prof_id>/delete", methods=["POST"])
@admin_required
def admin_professor_delete(prof_id):
    if prof_id == session.get("user_id"):
        flash("Você não pode excluir a si mesmo.", "error")
        return redirect(url_for("admin_professors"))
    db = get_db()
    db.execute("DELETE FROM users WHERE id=?", (prof_id,))
    db.commit()
    db.close()
    flash("Professor removido.", "success")
    return redirect(url_for("admin_professors"))

# --- Categories ---
@app.route("/admin/categories", methods=["GET", "POST"])
@admin_required
def admin_categories():
    db = get_db()
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if not name:
            flash("Nome da categoria é obrigatório.", "error")
        else:
            db.execute("INSERT INTO categories (name, order_idx) VALUES (?, (SELECT COALESCE(MAX(order_idx),0)+1 FROM categories))", (name,))
            db.commit()
            flash(f"Categoria '{name}' criada!", "success")
        db.close()
        return redirect(url_for("admin_categories"))

    cats = db.execute("SELECT * FROM categories ORDER BY order_idx").fetchall()
    db.close()
    return render_template("admin/categories.html", categories=cats)

@app.route("/admin/categories/<int:cat_id>/delete", methods=["POST"])
@admin_required
def admin_category_delete(cat_id):
    db = get_db()
    db.execute("DELETE FROM categories WHERE id=?", (cat_id,))
    db.commit()
    db.close()
    flash("Categoria removida.", "success")
    return redirect(url_for("admin_categories"))

# --- All Scores (admin view) ---
@app.route("/admin/scores")
@admin_required
def admin_scores():
    db = get_db()
    scores = db.execute("""
        SELECT s.id, s.points, s.note, s.created_at,
               t.name AS team_name, t.color,
               c.name AS category_name,
               u.name AS evaluator_name
        FROM scores s
        JOIN teams t ON t.id = s.team_id
        JOIN users u ON u.id = s.evaluator_id
        LEFT JOIN categories c ON c.id = s.category_id
        ORDER BY s.created_at DESC
    """).fetchall()
    db.close()
    return render_template("admin/scores.html", scores=scores)

@app.route("/admin/scores/<int:score_id>/delete", methods=["POST"])
@admin_required
def admin_score_delete(score_id):
    db = get_db()
    db.execute("DELETE FROM scores WHERE id=?", (score_id,))
    db.commit()
    db.close()
    flash("Pontuação excluída.", "success")
    return redirect(url_for("admin_scores"))

# ---------------------------------------------------------------------------
# Professor routes
# ---------------------------------------------------------------------------
@app.route("/professor")
@login_required
def professor_dashboard():
    if session.get("role") not in ("professor", "coordinator"):
        return redirect(url_for("admin_dashboard"))
    db = get_db()
    my_scores = db.execute("""
        SELECT s.id, s.points, s.note, s.created_at,
               t.name AS team_name, t.color,
               c.name AS category_name
        FROM scores s
        JOIN teams t ON t.id = s.team_id
        LEFT JOIN categories c ON c.id = s.category_id
        WHERE s.evaluator_id = ?
        ORDER BY s.created_at DESC
    """, (session["user_id"],)).fetchall()
    db.close()
    return render_template("professor/dashboard.html", my_scores=my_scores)

@app.route("/professor/evaluate", methods=["GET", "POST"])
@login_required
def professor_evaluate():
    if session.get("role") not in ("professor", "coordinator"):
        return redirect(url_for("admin_dashboard"))
    db = get_db()
    teams = db.execute("SELECT * FROM teams ORDER BY name").fetchall()
    cats  = db.execute("SELECT * FROM categories ORDER BY order_idx").fetchall()

    if request.method == "POST":
        team_id     = request.form.get("team_id", type=int)
        category_id = request.form.get("category_id", type=int) or None
        points      = request.form.get("points", type=float)
        note        = request.form.get("note", "").strip()

        if not team_id or points is None:
            flash("Selecione a equipe e informe a nota.", "error")
        elif points < 0 or points > 100:
            flash("A nota deve ser entre 0 e 100.", "error")
        else:
            try:
                db.execute(
                    "INSERT INTO scores (team_id, category_id, evaluator_id, points, note) VALUES (?, ?, ?, ?, ?)",
                    (team_id, category_id, session["user_id"], points, note),
                )
                db.commit()
                flash("Nota registrada!", "success")
            except sqlite3.IntegrityError:
                flash("Você já avaliou essa equipe nesta categoria. Exclua a avaliação antiga primeiro.", "error")
        db.close()
        return redirect(url_for("professor_evaluate"))

    db.close()
    return render_template("professor/evaluate.html", teams=teams, categories=cats)

# ---------------------------------------------------------------------------
# Backup / Export (admin) – Excel spreadsheet with all data
# ---------------------------------------------------------------------------
@app.route("/admin/backup")
@admin_required
def admin_backup():
    """Download every table as a styled Excel workbook (.xlsx)."""
    db = get_db()

    # --- fetch data ---
    teams = db.execute("""
        SELECT t.id, t.name, t.color,
               COALESCE(SUM(s.points), 0) AS total,
               COUNT(DISTINCT s.evaluator_id) AS evaluators
        FROM teams t
        LEFT JOIN scores s ON s.team_id = t.id
        GROUP BY t.id ORDER BY total DESC, t.name ASC
    """).fetchall()

    scores = db.execute("""
        SELECT t.name AS team_name, t.color,
               c.name AS category_name, s.points, s.note, s.created_at,
               u.name AS evaluator_name
        FROM scores s
        JOIN teams t ON t.id = s.team_id
        JOIN users u ON u.id = s.evaluator_id
        LEFT JOIN categories c ON c.id = s.category_id
        ORDER BY s.created_at DESC
    """).fetchall()

    all_teams = db.execute("SELECT id, name, color, created_at FROM teams ORDER BY name").fetchall()
    profs     = db.execute("SELECT name, email, created_at FROM users WHERE role='professor' ORDER BY name").fetchall()
    db.close()

    # --- styling helpers (UP palette) ---
    UP_RED   = "C8102E"
    UP_YELLOW = "F5C300"
    header_fill = PatternFill("solid", fgColor=UP_RED)
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font  = Font(bold=True, color=UP_RED, size=16)
    gold_fill   = PatternFill("solid", fgColor=UP_YELLOW)
    thin = Side(style="thin", color="E4E4E7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left", vertical="center")

    def style_header(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

    def autosize(ws, widths):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = w

    wb = openpyxl.Workbook()

    # ===== Sheet 1: Ranking =====
    ws = wb.active
    ws.title = "Ranking"
    ws.merge_cells("A1:D1")
    ws["A1"] = "🏆 Ranking Geral – Gincana UP 2026"
    ws["A1"].font = title_font
    ws.append([])
    hdr = ["Posição", "Equipe", "Pontuação", "Avaliadores"]
    ws.append(hdr)
    style_header(ws, 3, len(hdr))
    for i, t in enumerate(teams, start=1):
        ws.append([i, t["name"], float(t["total"]), t["evaluators"]])
        r = ws.max_row
        ws.cell(row=r, column=1).alignment = center
        ws.cell(row=r, column=3).alignment = center
        ws.cell(row=r, column=4).alignment = center
        if i == 1:
            ws.cell(row=r, column=1).fill = gold_fill
            ws.cell(row=r, column=1).font = Font(bold=True)
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = border
    autosize(ws, [10, 28, 14, 14])

    # ===== Sheet 2: Avaliações =====
    ws2 = wb.create_sheet("Avaliações")
    ws2.merge_cells("A1:G1")
    ws2["A1"] = "📝 Todas as Avaliações"
    ws2["A1"].font = title_font
    ws2.append([])
    hdr2 = ["Equipe", "Categoria", "Nota", "Avaliador", "Observação", "Data"]
    ws2.append(hdr2)
    style_header(ws2, 3, len(hdr2))
    for s in scores:
        ws2.append([
            s["team_name"],
            s["category_name"] or "Nota geral",
            float(s["points"]),
            s["evaluator_name"],
            s["note"] or "",
            (s["created_at"] or "")[:16],
        ])
        r = ws2.max_row
        for c in range(1, 7):
            ws2.cell(row=r, column=c).border = border
    autosize(ws2, [22, 16, 8, 22, 30, 18])

    # ===== Sheet 3: Equipes =====
    ws3 = wb.create_sheet("Equipes")
    ws3.merge_cells("A1:D1")
    ws3["A1"] = "👕 Equipes Cadastradas"
    ws3["A1"].font = title_font
    ws3.append([])
    hdr3 = ["ID", "Nome", "Cor", "Criado em"]
    ws3.append(hdr3)
    style_header(ws3, 3, len(hdr3))
    for t in all_teams:
        ws3.append([t["id"], t["name"], t["color"], (t["created_at"] or "")[:16]])
        r = ws3.max_row
        for c in range(1, 5):
            ws3.cell(row=r, column=c).border = border
    autosize(ws3, [6, 28, 12, 18])

    # ===== Sheet 4: Professores =====
    ws4 = wb.create_sheet("Professores")
    ws4.merge_cells("A1:C1")
    ws4["A1"] = "👨‍🏫 Professores Cadastrados"
    ws4["A1"].font = title_font
    ws4.append([])
    hdr4 = ["Nome", "E-mail", "Criado em"]
    ws4.append(hdr4)
    style_header(ws4, 3, len(hdr4))
    for p in profs:
        ws4.append([p["name"], p["email"], (p["created_at"] or "")[:16]])
        r = ws4.max_row
        for c in range(1, 4):
            ws4.cell(row=r, column=c).border = border
    autosize(ws4, [26, 30, 18])

    # freeze header rows
    for w in (ws, ws2, ws3, ws4):
        w.freeze_panes = "A4"

    # --- save to bytes ---
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"gincana_up_backup_{datetime.now():%Y%m%d_%H%M}.xlsx",
    )

# ---------------------------------------------------------------------------
# API – for ranking page live updates & AJAX
# ---------------------------------------------------------------------------
@app.route("/api/ranking")
def api_ranking():
    db = get_db()
    rows = db.execute("""
        SELECT t.id, t.name, t.color,
               COALESCE(SUM(s.points), 0) AS total,
               COUNT(DISTINCT s.evaluator_id) AS evaluator_count
        FROM teams t
        LEFT JOIN scores s ON s.team_id = t.id
        GROUP BY t.id
        ORDER BY total DESC, t.name ASC
    """).fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])

# ---------------------------------------------------------------------------
# Run
# ---------------------------------------------------------------------------
# Ensure the database schema (and default admin user) exist.
# Under gunicorn the module is imported, not run as __main__, so we
# initialize at import time. The operation is idempotent (IF NOT EXISTS).
init_db()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=DEBUG)
